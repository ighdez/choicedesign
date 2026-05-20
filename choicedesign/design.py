"""Classes for constructing efficient experimental designs.

The main user-facing class is :class:`EffDesign`, which combines initial
design generation, constrained random-swap optimisation, and optional blocking.

Typical workflow::

    design = EffDesign(X=[alt1_A, alt2_A], ncs=18)
    init   = design.gen_initdesign(cond=['alt1_A > alt2_A'], seed=42)
    result = design.optimise(init, V={1: V1, 2: V2}, time_lim=1)
    # result = (optimal_design, init_derr, final_derr, n_iter, utility_balance)
"""

# Import modules
import itertools
import pandas as pd
import numpy as np
import datetime
from typing import List

from choicedesign.expressions import Attribute
from choicedesign.algorithms import _swapalg, _rscalg, _federovalg
from choicedesign.criteria import MNLModel, _derr, _db_derr, _aerr, _cerr, _utility_balance
from choicedesign.utils import _blockgen, _parse_condition, _initdesign

# Efficient design
class EffDesign:
    """Efficient design for a discrete choice experiment.

    Combines initial design generation, random-swap optimisation, and
    optional blocking in a single object.

    Parameters
    ----------
    X : list[Attribute]
        Ordered list of :class:`~choicedesign.expressions.Attribute` objects
        that define the design columns and their discrete levels.
    ncs : int
        Number of choice situations (rows) in the design matrix.

    Examples
    --------
    >>> from choicedesign.design import EffDesign
    >>> from choicedesign.expressions import Attribute, Parameter
    >>> alt1_A = Attribute('alt1_A', [1, 2, 3])
    >>> alt2_A = Attribute('alt2_A', [1, 2, 3])
    >>> beta_A  = Parameter('beta_A', -0.1)
    >>> V = {1: beta_A * alt1_A, 2: beta_A * alt2_A}
    >>> design = EffDesign(X=[alt1_A, alt2_A], ncs=18)
    >>> init   = design.gen_initdesign(seed=42)
    >>> result = design.optimise(init, V=V, iter_lim=500)
    """

    # Init method
    def __init__(self, X: dict, ncs: int):

        # Define scalars
        self.N = ncs
        self.J = len(X)
 
        # Set names and levels
        self.names = [j.name for j in X]
        self.levs = [j.levels for j in X]

    # Generate initial design matrix
    def gen_initdesign(self,cond: list = None, seed: bool = None):
        """Generate initial design matrix

        It generates the initial design matrix. The user can define a set of
        conditions that must be satisfied.

        Parameters
        -------
        cond : list[str], optional
            List of conditions that the final design must hold. Each element
            is a string that contains a single condition. Supported forms:

            - Binary relation: ``'X > Y'`` (attribute vs attribute or value)
            - Conditional: ``'if X > a then Y < b'``
            - Compound (AND): ``'X > a & Y < b'``
            - Arithmetic expressions on either side: ``'(X + Y + Z) > 0'``,
              ``'if (X + Y) > 0 then P >= 0'``

            Arithmetic expressions support ``+``, ``-``, ``*``, ``/`` and
            parentheses with any mix of attribute names and numeric constants.
            By default None.
        seed : bool, None
            Random seed, by default None

        Returns
        -------
        init_design : pandas.DataFrame
            A Pandas DataFrame with the initial design matrix.
        """

        # Parse conditions into callables (validates attribute names immediately)
        if cond is not None:
            self.cond_callables = [_parse_condition(c, self.names) for c in cond]
        else:
            self.cond_callables = None

        # Set random seed if defined
        if seed is not None:
            np.random.seed(seed)

        # Generate initial design matrix
        init_design = _initdesign(levs=self.levs, ncs=self.N, cond=self.cond_callables)

        return pd.DataFrame(init_design,columns=self.names)

    # Optimise
    def optimise(self, init_design: pd.DataFrame, V: dict, model: str = 'mnl', algorithm: str = 'swap', criterion: str = 'd', cost_param=None, wtp_params=None, bayes_draws: int = None, iter_lim: int = None, noimprov_lim: int = None, time_lim: int = None, seed: int = None, verbose: bool = False):
        """Optimise the design using a random-search algorithm.

        Starts from an initial design and iteratively improves it according
        to the selected criterion and stopping rules. At least one stopping
        criterion (``iter_lim``, ``noimprov_lim``, or ``time_lim``) must be
        supplied.

        Parameters
        ----------
        init_design : pandas.DataFrame
            The initial design matrix, typically from :meth:`gen_initdesign`.
        V : dict
            A dictionary with the utility functions, keyed by alternative index.
            e.g. ``{1: V1, 2: V2}``
        model : str
            The base model for the efficient design, by default 'mnl'
        algorithm : str
            Optimisation algorithm to use. Options: ``'swap'`` (random swapping,
            default), ``'rsc'`` (random Relabelling, Swapping, Cycling), or
            ``'federov'`` (Modified Federov — tries all full-factorial candidates
            per row per iteration; slower per iteration but more systematic).
        criterion : str
            Optimality criterion: ``'d'`` (D-error, default), ``'a'`` (A-error),
            or ``'c'`` (C-error / WTP variance). When ``'c'``, ``cost_param``
            and ``wtp_params`` are required.
        cost_param : Parameter, optional
            The cost (denominator) parameter used to compute WTP ratios.
            Required when ``criterion='c'``.
        wtp_params : list[Parameter], optional
            Parameters whose WTP ratios are minimised. Required when
            ``criterion='c'``.
        bayes_draws : int, optional
            Number of Monte Carlo draws for Db-efficient (Bayesian) design.
            Only valid with ``criterion='d'``. When set, the optimizer minimises
            the expected D-error averaged over draws from the prior distributions
            of parameters that have ``prior_std`` defined.
        iter_lim : int, optional
            Number of iterations before the algorithm stops, by default None
        noimprov_lim : int, optional
            Number of iterations without improvement before the algorithm 
            stops, by default None
        time_lim : int, optional
            Time (in minutes) before the algorithm stops, by default None
        seed : int, optional
            Random seed, by default None
        verbose : bool, optional
            Whether status messages and progress are shown, by default False

        Returns
        -------
        optimal_design : pandas.DataFrame
            The final (optimal) design
        init_perf : float
            Criterion value of the initial design
        final_perf : float
            Criterion value of the final design
        final_iter : int
            Total number of iterations
        ubalance_ratio : float
            Utility balance ratio
        """
        # Set random seed if defined
        if seed is not None:
            np.random.seed(seed)

        # Keep original stopping-criteria values for the output report
        _iter_lim_orig     = iter_lim
        _noimprov_lim_orig = noimprov_lim
        _time_lim_orig     = time_lim

        # Set stopping criteria if defined
        if iter_lim is None:
            iter_lim = np.inf

        if noimprov_lim is None:
            noimprov_lim = np.inf

        if time_lim is None:
            time_lim = np.inf

        ############################################################
        ########## Step 1: Set initial design performance ##########
        ############################################################

        if verbose:
            print('Evaluating initial design')

        desmat = init_design

        if model == 'mnl':
            model_object = MNLModel(V)
        else:
            raise ValueError("Model name must be 'mnl'")

        if criterion == 'd':
            if bayes_draws is not None:
                rng = np.random.default_rng(seed)
                derr_fn = lambda design, model: _db_derr(design, model, bayes_draws, rng)
            else:
                derr_fn = _derr
        elif criterion == 'a':
            if bayes_draws is not None:
                raise ValueError("bayes_draws is only supported with criterion='d'")
            derr_fn = _aerr
        elif criterion == 'c':
            if cost_param is None or wtp_params is None:
                raise ValueError("criterion='c' requires both 'cost_param' and 'wtp_params'")
            if bayes_draws is not None:
                raise ValueError("bayes_draws is only supported with criterion='d'")
            derr_fn = lambda design, model: _cerr(design, model, cost_param, wtp_params)
        else:
            raise ValueError("criterion must be 'd', 'a', or 'c'")

        criterion_label = {'d': 'D-error', 'a': 'A-error', 'c': 'C-error'}[criterion]

        init_perf = derr_fn(desmat, model_object)

        ############################################################
        ############## Step 2: Initialize algorighm ################
        ############################################################

        # Execute optimisation algorithm
        if algorithm == 'swap':
            optimal_design, final_perf, final_iter, elapsed_time = _swapalg(
                desmat, model_object, init_perf, self.cond_callables, iter_lim, noimprov_lim, time_lim, derr_fn, criterion_label)
        elif algorithm == 'rsc':
            optimal_design, final_perf, final_iter, elapsed_time = _rscalg(
                desmat, model_object, init_perf, self.cond_callables, iter_lim, noimprov_lim, time_lim, derr_fn, criterion_label)
        elif algorithm == 'federov':
            optimal_design, final_perf, final_iter, elapsed_time = _federovalg(
                desmat, model_object, init_perf, self.cond_callables, iter_lim, noimprov_lim, time_lim, derr_fn, self.levs, criterion_label)
        else:
            raise ValueError("algorithm must be 'swap', 'rsc', or 'federov'")

        # Compute utility balance ratio
        ubalance_ratio = _utility_balance(pd.DataFrame(optimal_design, columns=self.names), model_object)

        ############################################################
        ############## Step 3: Arange final design #################
        ############################################################

        # Add CS column
        optimal_design = np.c_[np.arange(self.N)+1,optimal_design]

        # Generate blocks
        # if n_blocks is not None:
        #     if verbose:
        #         print('\nGenerating ' + str(n_blocks) + ' blocks...')
        #     blocksrow = _blockgen(optimal_design,n_blocks,self.N,1000)
        #     optimal_design = np.c_[optimal_design,blocksrow]

        # Create Pandas DataFrame
        # if n_blocks is not None:
        #     optimal_design = pd.DataFrame(optimal_design,columns=['CS'] + self.names + ['Block'])
        # else:
        optimal_design = pd.DataFrame(optimal_design,columns=['CS'] + self.names)

        # Store summary for export_output()
        self._last_output = {
            'timestamp':      datetime.datetime.now(),
            'ncs':            self.N,
            'attributes':     self.names,
            'model':          model,
            'algorithm':      algorithm,
            'criterion':      criterion,
            'criterion_label': criterion_label,
            'iter_lim':       _iter_lim_orig,
            'noimprov_lim':   _noimprov_lim_orig,
            'time_lim':       _time_lim_orig,
            'init_perf':      init_perf,
            'final_perf':     final_perf,
            'ubalance_ratio': ubalance_ratio,
            'final_iter':     final_iter,
            'elapsed_time':   elapsed_time,
        }

        # Return a summary if verbose is True
        if verbose:
            print('Optimization complete')
            print('Elapsed time: ' + str(datetime.timedelta(seconds=elapsed_time))[:7])
            print(f'{criterion_label} of initial design: ',round(init_perf,6))
            print(f'{criterion_label} of last stored design: ',round(final_perf,6))
            print('Utility Balance ratio: ',round(ubalance_ratio,2),'%')
            print('Algorithm iterations: ',final_iter)
            print('')

        # Return the optimal design
        return optimal_design, init_perf, final_perf, final_iter, ubalance_ratio

    # Generate blocks
    def gen_blocks(self, design: pd.DataFrame, n_blocks: int, n_iter: int = 1000):
        """Assign choice situations to blocks.

        Minimises the correlation between the block assignment and all
        attribute columns by evaluating ``n_iter`` random permutations and
        keeping the best one.

        Parameters
        ----------
        design : pandas.DataFrame
            Optimised design from :meth:`optimise` (must include a ``CS`` column).
        n_blocks : int
            Number of blocks.
        n_iter : int, optional
            Number of random permutations evaluated by the search, by default 1000.

        Returns
        -------
        design : pandas.DataFrame
            Design with an additional ``Block`` column.
        corr_list : list[float]
            History of best total absolute correlation found at each improvement.
        """
        blocksrow, corr_list = _blockgen(design,n_blocks,n_iter)
        design['Block'] = blocksrow

        return design, corr_list

    # Export design
    def export_design(self, design: pd.DataFrame, attr_names: dict, filepath: str, opt_out: bool = False, alt_names: list = None):
        """Export design to Excel in respondent-facing choice situation format.

        Writes one sheet per block (or a single sheet when the design has no
        ``Block`` column).  Inside each sheet, choice situations are stacked
        downward; each row is an attribute and each column is an alternative.

        Parameters
        ----------
        design : pd.DataFrame
            Design from :meth:`optimise` or :meth:`gen_blocks`.
        attr_names : dict
            Mapping from internal column names to display row labels.
            Columns that share the same display label appear in the same row
            (one per alternative column).
            Example: ``{'alt1_time': 'Travel time', 'alt2_time': 'Travel time',
            'alt1_cost': 'Cost', 'alt2_cost': 'Cost'}``
        filepath : str
            Destination path, e.g. ``'design.xlsx'``.
        opt_out : bool, optional
            Add an opt-out column with no attribute levels, by default False.
        alt_names : list[str], optional
            Custom headers for the alternative columns.  Defaults to
            ``['Alt 1', 'Alt 2', …]``.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise ImportError("openpyxl is required for export_design(). Install it with: pip install openpyxl")

        # Build ordered groups: display_name -> [col1, col2, ...]
        groups = {}
        for col, display in attr_names.items():
            if col not in design.columns:
                raise ValueError(f"Column '{col}' not found in design.")
            groups.setdefault(display, []).append(col)

        n_alts = max(len(cols) for cols in groups.values())

        if alt_names is None:
            alt_names = [f'Alt {i + 1}' for i in range(n_alts)]
        elif len(alt_names) != n_alts:
            raise ValueError(
                f"alt_names has {len(alt_names)} entries but {n_alts} alternatives were inferred."
            )

        col_headers = alt_names + (['Opt-out'] if opt_out else [])
        n_cols = 1 + len(col_headers)

        blocked = 'Block' in design.columns
        block_ids = sorted(design['Block'].unique()) if blocked else [None]

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        header_font = Font(bold=True)
        cs_fill = PatternFill(fill_type='solid', fgColor='D9E1F2')
        center = Alignment(horizontal='center', vertical='center')

        for block_id in block_ids:
            if blocked:
                sheet_name = f'Block {int(block_id)}'
                block_design = design[design['Block'] == block_id].reset_index(drop=True)
            else:
                sheet_name = 'Design'
                block_design = design.reset_index(drop=True)

            ws = wb.create_sheet(title=sheet_name)

            # Column header row
            ws.cell(row=1, column=1, value='Attribute').font = header_font
            for j, name in enumerate(col_headers):
                cell = ws.cell(row=1, column=j + 2, value=name)
                cell.font = header_font
                cell.alignment = center

            current_row = 2

            for _, row_data in block_design.iterrows():
                cs_val = int(row_data['CS'])

                # Choice-situation header (merged across all columns)
                ws.merge_cells(
                    start_row=current_row, start_column=1,
                    end_row=current_row, end_column=n_cols
                )
                cs_cell = ws.cell(row=current_row, column=1, value=f'Choice situation {cs_val}')
                cs_cell.font = header_font
                cs_cell.fill = cs_fill
                cs_cell.alignment = center
                current_row += 1

                # One row per display attribute
                for display_name, cols in groups.items():
                    ws.cell(row=current_row, column=1, value=display_name)
                    for j in range(n_alts):
                        if j < len(cols):
                            ws.cell(row=current_row, column=j + 2, value=row_data[cols[j]])
                    if opt_out:
                        ws.cell(row=current_row, column=n_cols, value='-')
                    current_row += 1

                current_row += 1  # blank row between choice situations

            # Approximate column widths
            for col in ws.columns:
                width = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max(width + 2, 12)

        wb.save(filepath)

    # Export optimisation output
    def export_output(self, filepath: str):
        """Save a plain-text optimisation summary to a file.

        The summary mirrors the information printed by :meth:`optimise` when
        ``verbose=True``, plus design configuration and stopping-criteria
        details.  :meth:`optimise` must have been called at least once before
        invoking this method.

        Parameters
        ----------
        filepath : str
            Destination path, e.g. ``'optimisation_summary.txt'``.

        Raises
        ------
        RuntimeError
            If :meth:`optimise` has not yet been called on this object.
        """
        if not hasattr(self, '_last_output'):
            raise RuntimeError(
                "No optimisation output available. Call optimise() first."
            )

        o = self._last_output
        crit   = o['criterion_label']
        sep    = '=' * 42
        subsep = '-' * 42

        def _fmt_lim(val, unit=''):
            if val is None:
                return '—'
            return f'{val}{unit}'

        improvement = (
            (o['init_perf'] - o['final_perf']) / o['init_perf'] * 100
            if o['init_perf'] != 0 else float('nan')
        )
        elapsed_str = str(datetime.timedelta(seconds=o['elapsed_time']))[:7]

        lines = [
            'ChoiceDesign — Optimisation Summary',
            sep,
            f"Generated  : {o['timestamp'].strftime('%Y-%m-%d %H:%M:%S')}",
            '',
            'Design configuration',
            subsep,
            f"Choice situations : {o['ncs']}",
            f"Attributes        : {', '.join(o['attributes'])}",
            f"Model             : {o['model']}",
            f"Algorithm         : {o['algorithm']}",
            f"Criterion         : {crit}",
            '',
            'Stopping criteria',
            subsep,
            f"Time limit (min)  : {_fmt_lim(o['time_lim'])}",
            f"Iteration limit   : {_fmt_lim(o['iter_lim'])}",
            f"No-improvement    : {_fmt_lim(o['noimprov_lim'])}",
            '',
            'Results',
            subsep,
            f"Initial {crit:<13}: {o['init_perf']:.6f}",
            f"Final {crit:<15}: {o['final_perf']:.6f}",
            f"Improvement       : {improvement:.1f} %",
            f"Utility balance   : {o['ubalance_ratio']:.2f} %",
            f"Iterations        : {o['final_iter']}",
            f"Elapsed time      : {elapsed_str}",
        ]

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines) + '\n')

    # Evaluate
    def evaluate(self, design: pd.DataFrame, V: dict, model: str = 'mnl', criterion: str = 'd', cost_param=None, wtp_params=None, bayes_draws: int = None, seed: int = None):
        """Evaluate design

        Evaluates a design stored in a Pandas data frame

        Parameters
        ----------
        design : pd.DataFrame
            Design to evaluate
        V : dict
            A dictionary with the utility function.
        model : str
            The base model for the efficient design, by default 'mnl'
        criterion : str
            Optimality criterion: ``'d'`` (D-error, default), ``'a'`` (A-error),
            or ``'c'`` (C-error / WTP variance). When ``'c'``, ``cost_param``
            and ``wtp_params`` are required.
        cost_param : Parameter, optional
            The cost (denominator) parameter. Required when ``criterion='c'``.
        wtp_params : list[Parameter], optional
            Parameters whose WTP variances are evaluated. Required when
            ``criterion='c'``.
        bayes_draws : int, optional
            Number of Monte Carlo draws for Db-error evaluation. Only valid
            with ``criterion='d'``.
        seed : int, optional
            Random seed for Bayesian draws, by default None

        Returns
        -------
        perf : float
            The criterion value of the design
        ubalance_ratio : float
            Utility balance ratio
        """
        # Drop CS column and Block (if present) from pandas dataframe
        desmat = design.drop('CS',axis=1)

        if 'Block' in desmat.columns:
            desmat = desmat.drop('Block',axis=1)

        if model == 'mnl':
            model_object = MNLModel(V)
        else:
            raise ValueError("Model name must be 'mnl'")

        # Evaluate the performance and utility balance of the design
        if criterion == 'd':
            if bayes_draws is not None:
                rng = np.random.default_rng(seed)
                perf = _db_derr(desmat, model_object, bayes_draws, rng)
            else:
                perf = _derr(desmat, model_object)
        elif criterion == 'a':
            if bayes_draws is not None:
                raise ValueError("bayes_draws is only supported with criterion='d'")
            perf = _aerr(desmat, model_object)
        elif criterion == 'c':
            if cost_param is None or wtp_params is None:
                raise ValueError("criterion='c' requires both 'cost_param' and 'wtp_params'")
            if bayes_draws is not None:
                raise ValueError("bayes_draws is only supported with criterion='d'")
            perf = _cerr(desmat, model_object, cost_param, wtp_params)
        else:
            raise ValueError("criterion must be 'd', 'a', or 'c'")
        ubalance_ratio = _utility_balance(desmat, model_object)

        # Return performance and utility balance
        return perf, ubalance_ratio
    
class FullFactDesign:
    """Full-factorial design covering all combinations of attribute levels.

    Parameters
    ----------
    X : list[Attribute]
        List of :class:`~choicedesign.expressions.Attribute` objects.

    Notes
    -----
    This class is provided for completeness. Its dependency ``pyDOE2`` is
    incompatible with Python 3.12+ and is not installed by default.
    """

    def __init__(self, X: list):
        self.names = [j.name for j in X]
        self.levs = [j.levels for j in X]

    def gen_design(self):
        """Generate full-factorial design matrix

        Returns
        -------
        design : pandas.DataFrame
            A Pandas DataFrame with all combinations of attribute levels.
        """
        rows = list(itertools.product(*self.levs))
        return pd.DataFrame(rows, columns=self.names)

#         return init_design